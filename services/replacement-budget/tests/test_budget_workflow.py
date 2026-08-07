import io
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from app import app, supabase


class BudgetWorkflowTests(unittest.TestCase):
    def setUp(self):
        app.config.update(TESTING=True)
        self.client = app.test_client()

    def _select_schools(self):
        response = self.client.post(
            "/budget/schools",
            data={"school_id": ["school-a", "school-c"]},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.location.endswith("/budget/fixtures"))

    def _select_fixtures(self):
        self._select_schools()
        with self.client.session_transaction() as session:
            selected = list(session["budget_state"]["selected_fixtures"])
        response = self.client.post(
            "/budget/fixtures",
            data={"fixture_id": selected},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.location.endswith("/budget/replacements"))
        return selected

    def _build_budget(self):
        selected = self._select_fixtures()
        data = {"labor_cost": "2500.00"}
        for fixture_id in selected:
            data[f"part_{fixture_id}"] = "Water Fountain"
            data[f"cost_{fixture_id}"] = "1750.00"
        response = self.client.post("/budget/replacements", data=data)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.location.endswith("/budget/review"))
        return selected

    def test_first_step_is_district_specific_and_has_placeholder_schools(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Estimate Budget for Lead Remediation", response.data)
        self.assertIn(b"Choose schools to include", response.data)
        self.assertNotIn(b"North Valley School District", response.data)
        self.assertIn(b"School A", response.data)
        self.assertIn(b"School B", response.data)
        self.assertIn(b"School C", response.data)
        self.assertIn(b"Select all", response.data)
        self.assertIn(b"Clear", response.data)
        self.assertIn(b"school-dropdown", response.data)
        self.assertIn(b'placeholder="Search by school name"', response.data)
        self.assertIn(b"No schools selected", response.data)
        self.assertNotIn(b"101 Cedar Avenue", response.data)
        self.assertNotIn(b"The district is already matched", response.data)
        self.assertNotIn(b"Remediation scope", response.data)
        self.assertIn(b"Select school(s)", response.data)
        self.assertIn(b"Select fixture(s)", response.data)
        self.assertNotIn(b"<strong>Select schools</strong>", response.data)
        self.assertNotIn(b"<strong>Select fixtures</strong>", response.data)

    def test_aquatrack_navigation_and_auth_launch_are_available(self):
        response = self.client.get("/auth/launch")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"AquaTrack Replacement Budget", response.data)

        response = self.client.get("/")
        self.assertIn(b"Fixture Inventory", response.data)
        self.assertIn(b"Communication", response.data)
        self.assertIn(b"Replacement Budget", response.data)

    def test_school_dropdown_shows_selected_school_names(self):
        self._select_schools()
        response = self.client.get("/budget/schools")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"School A \xc2\xb7 School C", response.data)
        self.assertIn(b"2</strong> selected", response.data)

    def test_new_aquatrack_launch_starts_with_no_selected_schools(self):
        self._select_schools()
        with self.client.session_transaction() as current_session:
            self.assertEqual(
                current_session["budget_state"]["selected_schools"],
                ["school-a", "school-c"],
            )

        with patch.object(supabase, "verify_user", return_value={"id": "user-1"}):
            response = self.client.post(
                "/auth/session",
                json={"access_token": "valid-token"},
            )
        self.assertEqual(response.status_code, 200)

        with self.client.session_transaction() as current_session:
            self.assertNotIn("budget_state", current_session)
        response = self.client.get("/budget/schools")
        self.assertIn(b"No schools selected", response.data)
        self.assertNotIn(b'name="school_id" value="school-a" checked', response.data)

    def test_fixtures_default_to_all_results_above_five_ppb(self):
        self._select_schools()
        response = self.client.get("/budget/fixtures")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Deselect all", response.data)
        self.assertIn(b"data-toggle-eligible", response.data)
        self.assertNotIn(b"Select all eligible", response.data)
        self.assertNotIn(b">Clear<", response.data)
        self.assertNotIn(b"> Back<", response.data)
        self.assertNotIn(b"Results are sorted from highest to lowest", response.data)
        self.assertNotIn(b"Results above 5 ppb", response.data)
        html = response.data.decode("utf-8")
        self.assertLess(html.index("A-101"), html.index("C-301"))
        self.assertLess(html.index("C-301"), html.index("A-106"))
        with self.client.session_transaction() as session:
            state = session["budget_state"]
        self.assertEqual(
            state["selected_fixtures"],
            ["A-101", "C-301", "A-102", "C-302", "A-103", "C-303", "A-104", "C-304"],
        )
        self.assertNotIn("A-105", state["selected_fixtures"])
        self.assertNotIn("A-106", state["selected_fixtures"])

    def test_replacement_defaults_and_vendor_reference_list(self):
        self._select_fixtures()
        response = self.client.get("/budget/replacements")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Vendors nearby", response.data)
        self.assertNotIn(b"Placeholder vendors", response.data)
        self.assertIn(b"ClearFlow School Plumbing", response.data)
        self.assertNotIn(b"preferred_vendor_id", response.data)
        self.assertIn(b'value="Water Fountain"', response.data)
        self.assertIn(b'value="1500.00"', response.data)
        self.assertIn(b'data-default-cost="600"', response.data)
        self.assertIn(b'data-labor-cost', response.data)
        self.assertIn(b"Review budget", response.data)
        self.assertNotIn(b"> Back<", response.data)
        self.assertNotIn(b"Each fixture starts with a same-type replacement", response.data)
        self.assertNotIn(b"Parts, labor &amp; vendors", response.data)

    def test_review_and_excel_export_include_custom_costs_and_labor(self):
        selected = self._build_budget()
        response = self.client.get("/budget/review")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Review budget and export", response.data)
        self.assertNotIn(b"Review the replacement budget", response.data)
        self.assertNotIn(b"Confirm the scope and estimates", response.data)
        self.assertNotIn(b"Ready to export", response.data)
        self.assertNotIn(b"Back to costs", response.data)
        self.assertNotIn(b"Export district budget", response.data)
        self.assertIn(b"Generate Excel workbook", response.data)
        expected_material = len(selected) * 1750
        expected_total = expected_material + 2500
        self.assertIn(f"${expected_material:.2f}".encode(), response.data)
        self.assertIn(f"${expected_total:.2f}".encode(), response.data)

        export = self.client.post("/budget/export")
        self.assertEqual(export.status_code, 200)
        self.assertEqual(
            export.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(io.BytesIO(export.data), data_only=True)
        self.assertEqual(workbook.sheetnames, ["Budget Summary", "Budget Detail"])
        summary = workbook["Budget Summary"]
        summary_values = {
            summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value
            for row in range(3, 14)
        }
        self.assertEqual(summary_values["District"], "North Valley School District")
        self.assertEqual(summary_values["Material subtotal"], expected_material)
        self.assertEqual(summary_values["Labor cost"], 2500)
        self.assertEqual(summary_values["Total estimated budget"], expected_total)
        self.assertEqual(workbook["Budget Detail"].max_row, len(selected) + 1)

    def test_later_steps_require_prior_completion(self):
        response = self.client.get("/budget/replacements")
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.location.endswith("/budget/schools"))
        response = self.client.get("/budget/review")
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.location.endswith("/budget/schools"))


if __name__ == "__main__":
    unittest.main()
