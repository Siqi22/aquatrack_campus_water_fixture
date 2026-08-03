"""Budget workflow data, calculations, and Excel export helpers.

The fixture records in this module are intentionally placeholders.  The UI and
route layer consume them through small helper functions so they can later be
replaced by Supabase queries without changing the four-step workflow.
"""

from __future__ import annotations

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.query_engine import CONTAMINATION_THRESHOLD_PPB, UNIT_COSTS


DISTRICT_NAME = "North Valley School District"

SCHOOLS = [
    {
        "id": "school-a",
        "name": "School A",
        "type": "Elementary School",
        "address": "101 Cedar Avenue",
    },
    {
        "id": "school-b",
        "name": "School B",
        "type": "Middle School",
        "address": "240 Valley Road",
    },
    {
        "id": "school-c",
        "name": "School C",
        "type": "High School",
        "address": "880 Northview Drive",
    },
]

VENDORS = [
    {
        "id": "clearflow",
        "name": "ClearFlow School Plumbing",
        "city": "North Valley",
        "distance_miles": 4,
        "phone": "(509) 555-0142",
        "email": "quotes@clearflow.example",
        "services": "Fixture replacement · drinking fountains · school plumbing",
    },
    {
        "id": "evergreen-water",
        "name": "Evergreen Water Systems",
        "city": "Cedar Junction",
        "distance_miles": 11,
        "phone": "(509) 555-0188",
        "email": "estimating@evergreenwater.example",
        "services": "Bottle refill stations · filtration · installation",
    },
    {
        "id": "inland-fixture",
        "name": "Inland Fixture & Mechanical",
        "city": "Pine Ridge",
        "distance_miles": 18,
        "phone": "(509) 555-0165",
        "email": "schoolquotes@inlandfixture.example",
        "services": "Commercial sinks · coolers · labor estimates",
    },
]

# Covers the full requested 0-50 ppb range and multiple part types.  Values are
# deterministic so UI tests and future Supabase mapping can rely on stable IDs.
FIXTURES = [
    {"id": "A-101", "school_id": "school-a", "location": "Cafeteria, north wall", "fixture_type": "Water Fountain", "lead_ppb": 50.0, "sample_date": "2026-04-12"},
    {"id": "A-102", "school_id": "school-a", "location": "Room 12 classroom sink", "fixture_type": "Tap/Sink", "lead_ppb": 32.4, "sample_date": "2026-04-12"},
    {"id": "A-103", "school_id": "school-a", "location": "Main hallway", "fixture_type": "Bottle Refill Station", "lead_ppb": 17.8, "sample_date": "2026-04-12"},
    {"id": "A-104", "school_id": "school-a", "location": "Nurse office sink", "fixture_type": "Tap/Sink", "lead_ppb": 8.2, "sample_date": "2026-04-12"},
    {"id": "A-105", "school_id": "school-a", "location": "Library fountain", "fixture_type": "Water Fountain", "lead_ppb": 5.0, "sample_date": "2026-04-12"},
    {"id": "A-106", "school_id": "school-a", "location": "Staff lounge sink", "fixture_type": "Tap/Sink", "lead_ppb": 0.0, "sample_date": "2026-04-12"},
    {"id": "B-201", "school_id": "school-b", "location": "Science lab 1 sink", "fixture_type": "Tap/Sink", "lead_ppb": 44.1, "sample_date": "2026-04-18"},
    {"id": "B-202", "school_id": "school-b", "location": "Gym entrance", "fixture_type": "Water Fountain", "lead_ppb": 26.0, "sample_date": "2026-04-18"},
    {"id": "B-203", "school_id": "school-b", "location": "Second-floor commons", "fixture_type": "Bottle Refill Station", "lead_ppb": 12.6, "sample_date": "2026-04-18"},
    {"id": "B-204", "school_id": "school-b", "location": "Art room sink", "fixture_type": "Tap/Sink", "lead_ppb": 6.1, "sample_date": "2026-04-18"},
    {"id": "B-205", "school_id": "school-b", "location": "Kitchen prep sink", "fixture_type": "Sprayer/Hose", "lead_ppb": 4.2, "sample_date": "2026-04-18"},
    {"id": "B-206", "school_id": "school-b", "location": "Counseling office fountain", "fixture_type": "Water Cooler", "lead_ppb": 1.4, "sample_date": "2026-04-18"},
    {"id": "C-301", "school_id": "school-c", "location": "Culinary room pot filler", "fixture_type": "Pot Filler", "lead_ppb": 38.5, "sample_date": "2026-04-25"},
    {"id": "C-302", "school_id": "school-c", "location": "Athletics hallway", "fixture_type": "Water Fountain", "lead_ppb": 21.7, "sample_date": "2026-04-25"},
    {"id": "C-303", "school_id": "school-c", "location": "Student commons", "fixture_type": "Bottle Refill Station", "lead_ppb": 9.8, "sample_date": "2026-04-25"},
    {"id": "C-304", "school_id": "school-c", "location": "Chemistry lab sink", "fixture_type": "Tap/Sink", "lead_ppb": 5.6, "sample_date": "2026-04-25"},
    {"id": "C-305", "school_id": "school-c", "location": "Faculty lounge ice machine", "fixture_type": "Ice Machine/Fridge", "lead_ppb": 3.1, "sample_date": "2026-04-25"},
    {"id": "C-306", "school_id": "school-c", "location": "Auditorium lobby fountain", "fixture_type": "Water Fountain", "lead_ppb": 0.7, "sample_date": "2026-04-25"},
]

SCHOOL_BY_ID = {school["id"]: school for school in SCHOOLS}
FIXTURE_BY_ID = {fixture["id"]: fixture for fixture in FIXTURES}


def replacement_options() -> list[dict[str, int | str]]:
    """Return the existing system's replacement catalog in UI order."""
    preferred_order = [
        "Tap/Sink",
        "Water Fountain",
        "Bottle Refill Station",
        "Water Cooler",
        "Ice Machine/Fridge",
        "Pot Filler",
        "Sprayer/Hose",
        "Other",
    ]
    return [
        {"name": name, "cost": int(UNIT_COSTS[name])}
        for name in preferred_order
        if name in UNIT_COSTS
    ]


def fixtures_for_schools(
    school_ids: list[str],
    fixtures_source: list[dict] | None = None,
    school_by_id: dict[str, dict] | None = None,
) -> list[dict]:
    """Return fixtures for selected schools, highest ppb first.

    The optional sources let AquaTrack provide the authenticated user's live
    Supabase inventory while preserving the original placeholder workflow for
    standalone development and regression tests.
    """
    fixtures_source = FIXTURES if fixtures_source is None else fixtures_source
    school_by_id = SCHOOL_BY_ID if school_by_id is None else school_by_id
    selected = set(school_ids)
    fixtures = []
    for fixture in fixtures_source:
        if fixture["school_id"] not in selected:
            continue
        enriched = dict(fixture)
        enriched.setdefault("display_id", fixture["id"])
        school = school_by_id.get(fixture["school_id"], {})
        enriched["school_name"] = school.get("name", "School")
        enriched["eligible"] = fixture["lead_ppb"] > CONTAMINATION_THRESHOLD_PPB
        fixtures.append(enriched)
    return sorted(fixtures, key=lambda row: (-row["lead_ppb"], row["id"]))


def normalize_school_ids(values: list[str], schools: list[dict] | None = None) -> list[str]:
    schools = SCHOOLS if schools is None else schools
    valid = {school["id"] for school in schools}
    return [school["id"] for school in schools if school["id"] in values and school["id"] in valid]


def normalize_fixture_ids(
    values: list[str],
    school_ids: list[str],
    fixtures_source: list[dict] | None = None,
    school_by_id: dict[str, dict] | None = None,
) -> list[str]:
    available = fixtures_for_schools(school_ids, fixtures_source, school_by_id)
    eligible = {
        fixture["id"]
        for fixture in available
        if fixture["eligible"]
    }
    return [fixture["id"] for fixture in available if fixture["id"] in values and fixture["id"] in eligible]


def default_replacement(fixture: dict) -> dict[str, int | str]:
    part = fixture["fixture_type"] if fixture["fixture_type"] in UNIT_COSTS else "Other"
    return {"part": part, "unit_cost": int(UNIT_COSTS.get(part, UNIT_COSTS["Other"]))}


def build_budget_lines(
    state: dict,
    fixture_by_id: dict[str, dict] | None = None,
    school_by_id: dict[str, dict] | None = None,
) -> list[dict]:
    fixture_by_id = FIXTURE_BY_ID if fixture_by_id is None else fixture_by_id
    school_by_id = SCHOOL_BY_ID if school_by_id is None else school_by_id
    lines = []
    replacements = state.get("replacements", {})
    for fixture_id in state.get("selected_fixtures", []):
        fixture = fixture_by_id.get(fixture_id)
        if not fixture:
            continue
        replacement = replacements.get(fixture_id) or default_replacement(fixture)
        line = dict(fixture)
        line.setdefault("display_id", fixture["id"])
        line["school_name"] = school_by_id.get(fixture["school_id"], {}).get("name", "School")
        line["replacement_part"] = replacement["part"]
        line["unit_cost"] = float(replacement["unit_cost"])
        lines.append(line)
    return sorted(lines, key=lambda row: (-row["lead_ppb"], row["id"]))


def budget_totals(
    state: dict,
    fixture_by_id: dict[str, dict] | None = None,
    school_by_id: dict[str, dict] | None = None,
) -> dict[str, float]:
    material_cost = sum(
        line["unit_cost"]
        for line in build_budget_lines(state, fixture_by_id, school_by_id)
    )
    labor_cost = max(0.0, float(state.get("labor_cost", 0) or 0))
    return {
        "material_cost": material_cost,
        "labor_cost": labor_cost,
        "total_cost": material_cost + labor_cost,
    }


def generate_budget_xlsx(
    state: dict,
    district_name: str = DISTRICT_NAME,
    fixture_by_id: dict[str, dict] | None = None,
    school_by_id: dict[str, dict] | None = None,
) -> tuple[bytes, str]:
    """Generate a review-ready workbook for the current budget state."""
    fixture_by_id = FIXTURE_BY_ID if fixture_by_id is None else fixture_by_id
    school_by_id = SCHOOL_BY_ID if school_by_id is None else school_by_id
    lines = build_budget_lines(state, fixture_by_id, school_by_id)
    totals = budget_totals(state, fixture_by_id, school_by_id)
    selected_school_names = [
        school_by_id[school_id]["name"]
        for school_id in state.get("selected_schools", [])
        if school_id in school_by_id
    ]

    wb = Workbook()
    summary = wb.active
    summary.title = "Budget Summary"
    detail = wb.create_sheet("Budget Detail")

    teal = "168899"
    dark = "1C2A33"
    pale = "E8F5F7"
    border_color = "DDE3E7"
    thin = Side(style="thin", color=border_color)

    summary.merge_cells("A1:D1")
    summary["A1"] = f"{district_name} — Lead Remediation Budget"
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=teal)
    summary["A1"].alignment = Alignment(horizontal="left", vertical="center")
    summary.row_dimensions[1].height = 30

    summary_rows = [
        ("District", district_name),
        ("Generated", datetime.now().strftime("%Y-%m-%d")),
        ("Lead threshold", f"> {CONTAMINATION_THRESHOLD_PPB} ppb"),
        ("Selected schools", ", ".join(selected_school_names)),
        ("Fixtures to replace", len(lines)),
        ("Material subtotal", totals["material_cost"]),
        ("Labor cost", totals["labor_cost"]),
        ("Total estimated budget", totals["total_cost"]),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(row=row_index, column=1, value=label)
        summary.cell(row=row_index, column=2, value=value)
        summary.cell(row=row_index, column=1).font = Font(bold=True, color=dark)
        summary.cell(row=row_index, column=1).fill = PatternFill("solid", fgColor=pale)
        for col in (1, 2):
            summary.cell(row=row_index, column=col).border = Border(bottom=thin)
            summary.cell(row=row_index, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        if label in {"Material subtotal", "Labor cost", "Total estimated budget"}:
            summary.cell(row=row_index, column=2).number_format = '"$"#,##0.00'
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 54

    headers = [
        "School",
        "Fixture ID",
        "Location",
        "Existing Fixture Type",
        "Lead Result (ppb)",
        "Replacement Part",
        "Estimated Unit Cost",
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = detail.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=teal)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:G{max(2, len(lines) + 1)}"

    for row_index, line in enumerate(lines, start=2):
        values = [
            line["school_name"],
            line["display_id"],
            line["location"],
            line["fixture_type"],
            line["lead_ppb"],
            line["replacement_part"],
            line["unit_cost"],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = detail.cell(row=row_index, column=col_index, value=value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F6FAFB")
        detail.cell(row=row_index, column=5).number_format = "0"
        detail.cell(row=row_index, column=7).number_format = '"$"#,##0.00'

    widths = [20, 13, 34, 24, 18, 26, 22]
    for col_index, width in enumerate(widths, start=1):
        detail.column_dimensions[get_column_letter(col_index)].width = width
    detail.row_dimensions[1].height = 32

    output = io.BytesIO()
    wb.save(output)
    safe_district = re.sub(r"[^A-Za-z0-9]+", "_", district_name).strip("_")
    return output.getvalue(), f"{safe_district}_remediation_budget.xlsx"
