"""
csv_generator.py
----------------
Generates a downloadable .xlsx report for a single school district,
using the result dict returned by query_engine.search_district().

Output: two-tab Excel workbook
    Tab 1 "Grant Application" — school name, fixture type, fixtures above 5ppb,
                                grade level, total estimated cost (merged cells)
    Tab 2 "Full Detail"       — all columns including location, ppb, year, unit cost

Usage:
    from backend.query_engine  import search_district
    from backend.csv_generator import generate_district_xlsx

    result = search_district("Kent")
    path   = generate_district_xlsx(result, output_dir="exports/")
"""

import io
import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Cost config ───────────────────────────────────────────────────────────────

UNIT_COSTS = {
    "Tap/Sink":               600,
    "Water Fountain":        1500,
    "Bottle Refill Station": 1500,
    "Water Cooler":           800,
    "Ice Machine/Fridge":     800,
    "Pot Filler":             600,
    "Sprayer/Hose":           400,
    "Other":                  600,
}

CONTAMINATION_THRESHOLD_PPB = 5


# ── Helpers ───────────────────────────────────────────────────────────────────

def _infer_grade_level(school_name: str) -> str:
    name = school_name.lower()
    if any(kw in name for kw in ["k-12", "k12"]):
        return "K-12"
    if any(kw in name for kw in ["k-8", "k8"]):
        return "K-8"
    if any(kw in name for kw in ["headstart", "head start", "preschool",
                                   "eceap", "early learning", "early childhood",
                                   "kindergarten"]):
        return "Early Learning"
    if any(kw in name for kw in ["elementary", "primary", "elem "]):
        return "Elementary"
    if name.split()[-1] in ("elem", "elem.", "es"):
        return "Elementary"
    if any(kw in name for kw in ["middle", "intermediate"]):
        return "Middle"
    if any(kw in name for kw in ["high", "junior", "jr/sr", "jr. high"]):
        return "High"
    return "Unknown"


def _clean_location(loc: str) -> str:
    """Strip trailing semicolons and extra whitespace from a location string."""
    if not isinstance(loc, str):
        return ""
    return loc.strip().strip(";").strip()


def _safe_filename(name: str) -> str:
    return re.sub(r"[^\w\s-]", "", name).strip().replace(" ", "_")


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths: dict):
    """widths = { "A": 30, "B": 20, ... }"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_header_row(ws, row: int, headers: list[str], fill_hex: str):
    fill = _header_fill(fill_hex)
    bold = Font(bold=True, color="FFFFFF")
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


# ── Build the fixture detail dataframe (one row per location) ─────────────────

def _build_detail_rows(fixture_rows: pd.DataFrame) -> pd.DataFrame:
    """
    Returns a DataFrame where each row is one contaminated fixture location.
    Columns: school_name, fixture_type, location, avg_lead_ppb,
             year_sampled, grade_level, unit_cost, total_cost_for_group,
             fixtures_above_5ppb (count for the school+type group)
    """
    if fixture_rows.empty:
        return pd.DataFrame()

    rows = []
    groups = fixture_rows.groupby(
        ["school_name", "fixture_type", "DOH_testing_round"], sort=True
    )

    for (school, ftype, year), grp in groups:
        count     = len(grp)
        unit_cost = UNIT_COSTS.get(ftype, UNIT_COSTS["Other"])
        total     = unit_cost * count
        grade     = _infer_grade_level(school)

        for _, fixture_row in grp.iterrows():
            loc = _clean_location(str(fixture_row.get("fixture_housing_location", "")))
            rows.append({
                "school_name":           school,
                "fixture_type":          ftype,
                "fixture_location":      loc,
                "fixtures_above_5ppb":   count,
                "avg_lead_ppb":          round(float(fixture_row.get("mean_lead_result_ppb", 0)), 1),
                "year_sampled":          int(year),
                "grade_level":           grade,
                "unit_replacement_cost": unit_cost,
                "total_estimated_cost":  total,
            })

    return pd.DataFrame(rows)


# ── Tab 1: Grant Application ──────────────────────────────────────────────────

GRANT_HEADERS = [
    "School Name",
    "Fixture Type",
    "Fixtures Above 5 ppb",
    "Grade Level",
    "Total Estimated Cost",
]

GRANT_COLS = ["school_name", "fixture_type", "fixtures_above_5ppb",
              "grade_level", "total_estimated_cost"]

# Columns that get merged vertically when values repeat within a school+fixture group
# Col A (school_name) merges across all location rows for that school+fixture
GRANT_MERGE_COLS = {1: "school_name"}   # col index → field name


def _write_grant_tab(ws, detail_df: pd.DataFrame, district_name: str, summary: dict):
    ws.title = "Grant Application"

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"{district_name} School District — Lead Remediation Grant Summary"
    title_cell.font  = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value = (
        f"Generated {datetime.now().strftime('%Y-%m-%d')}  |  "
        f"Data: WA DOH Lead in School Drinking Water Program  |  "
        f"Threshold: {CONTAMINATION_THRESHOLD_PPB} ppb"
    )
    sub.font      = Font(italic=True, size=9, color="666666")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    # ── Headers ───────────────────────────────────────────────────────────────
    _write_header_row(ws, 3, GRANT_HEADERS, "2E75B6")
    ws.row_dimensions[3].height = 18

    if detail_df.empty:
        ws.cell(row=4, column=1, value="No contaminated fixtures found for this district.")
        return

    # Grant tab shows one row per school+fixture_type group (not per location)
    # Deduplicate to group level
    grant_df = (
        detail_df
        .groupby(["school_name", "fixture_type", "grade_level"], sort=True)
        .agg(
            fixtures_above_5ppb  = ("fixtures_above_5ppb", "first"),
            total_estimated_cost = ("total_estimated_cost", "first"),
        )
        .reset_index()
    )[GRANT_COLS]

    current_row = 4
    alt_fill    = _header_fill("EBF3FB")   # light blue alternating rows

    for school, school_grp in grant_df.groupby("school_name", sort=True):
        school_start = current_row

        for _, row_data in school_grp.iterrows():
            for col_idx, field in enumerate(GRANT_COLS, 1):
                val  = row_data[field]
                cell = ws.cell(row=current_row, column=col_idx)

                if field == "total_estimated_cost" and isinstance(val, (int, float)):
                    cell.value        = val
                    cell.number_format = '"$"#,##0'
                else:
                    cell.value = val

                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border    = _thin_border()
                if current_row % 2 == 0:
                    cell.fill = alt_fill

            current_row += 1

        # Merge school name (col A) across all rows for this school
        if current_row - school_start > 1:
            ws.merge_cells(
                start_row=school_start, start_column=1,
                end_row=current_row - 1,   end_column=1
            )
            ws["A" + str(school_start)].alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

    # ── District total row ────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 16
    total_cost = summary.get("remediation_cost_total", 0)

    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row,   end_column=4
    )
    label_cell = ws.cell(row=current_row, column=1,
                         value="TOTAL DISTRICT REMEDIATION COST ESTIMATE")
    label_cell.font      = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    label_cell.fill      = _header_fill("D6E4F0")

    cost_cell = ws.cell(row=current_row, column=5, value=total_cost)
    cost_cell.font          = Font(bold=True)
    cost_cell.number_format = '"$"#,##0'
    cost_cell.alignment     = Alignment(horizontal="center", vertical="center")
    cost_cell.fill          = _header_fill("D6E4F0")
    cost_cell.border        = _thin_border()

    # ── Column widths ─────────────────────────────────────────────────────────
    _set_col_widths(ws, {"A": 36, "B": 20, "C": 22, "D": 16, "E": 22})


# ── Tab 2: Full Detail ────────────────────────────────────────────────────────

FULL_HEADERS = [
    "School Name",
    "Fixture Type",
    "Fixture Location",
    "Fixtures Above 5 ppb",
    "Avg Lead (ppb)",
    "Year Sampled",
    "Grade Level",
    "Unit Replacement Cost",
    "Total Estimated Cost",
]

FULL_COLS = [
    "school_name", "fixture_type", "fixture_location",
    "fixtures_above_5ppb", "avg_lead_ppb", "year_sampled",
    "grade_level", "unit_replacement_cost", "total_estimated_cost",
]

# Which columns merge vertically within a school+fixture_type group
# col index → field that should be blank after the first location row
FULL_MERGE_COLS = {1, 2, 4, 5, 6, 7, 8, 9}   # everything except fixture_location (col 3)


def _write_full_tab(ws, detail_df: pd.DataFrame, district_name: str, summary: dict):
    ws.title = "Full Detail"

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"{district_name} School District — Full Lead Testing Detail"
    title_cell.font  = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Headers ───────────────────────────────────────────────────────────────
    _write_header_row(ws, 2, FULL_HEADERS, "375623")   # dark green
    ws.row_dimensions[2].height = 18

    if detail_df.empty:
        ws.cell(row=3, column=1, value="No contaminated fixtures found for this district.")
        return

    current_row = 3
    alt_fill    = _header_fill("EFF5EE")   # light green alternating

    for (school, ftype), grp in detail_df.groupby(
        ["school_name", "fixture_type"], sort=True
    ):
        group_start = current_row

        for loc_idx, (_, row_data) in enumerate(grp.iterrows()):
            for col_idx, field in enumerate(FULL_COLS, 1):
                cell = ws.cell(row=current_row, column=col_idx)

                # Only write repeated group-level values on the first location row
                if col_idx in FULL_MERGE_COLS and loc_idx > 0:
                    cell.value = None
                else:
                    val = row_data[field]
                    if field in ("unit_replacement_cost", "total_estimated_cost") \
                            and isinstance(val, (int, float)):
                        cell.value        = val
                        cell.number_format = '"$"#,##0'
                    else:
                        cell.value = val

                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border    = _thin_border()
                if current_row % 2 == 0:
                    cell.fill = alt_fill

            current_row += 1

        # Merge group-level columns across all location rows
        if current_row - group_start > 1:
            for col_idx in FULL_MERGE_COLS:
                ws.merge_cells(
                    start_row=group_start, start_column=col_idx,
                    end_row=current_row - 1, end_column=col_idx
                )
                merged_cell = ws.cell(row=group_start, column=col_idx)
                merged_cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

    # ── Summary block at bottom ───────────────────────────────────────────────
    current_row += 1   # blank gap

    summary_items = [
        ("District",                    district_name),
        ("Testing Round(s)",            ", ".join(str(r) for r in summary["testing_rounds"])),
        ("Total Schools",               summary["schools_total"]),
        ("Schools with Contamination",  summary["schools_contaminated"]),
        ("Total Fixtures Tested",       summary["fixtures_tested"]),
        ("Fixtures Above 5 ppb",        summary["fixtures_contaminated"]),
        ("% Fixtures Contaminated",     f"{summary['pct_fixtures_contaminated']}%"),
    ]
    for ft, cost in summary["remediation_cost_by_type"].items():
        count = summary["fixtures_above_by_type"].get(ft, 0)
        unit  = UNIT_COSTS.get(ft, UNIT_COSTS["Other"])
        summary_items.append((f"  {ft} ({count} × ${unit:,})", f"${cost:,}"))
    summary_items.append(("TOTAL REMEDIATION COST", f"${summary['remediation_cost_total']:,}"))
    summary_items.append(("Cost Note", "Material replacement only. Labor/inspection not included."))
    summary_items.append(("Data Source", "WA DOH Lead in School Drinking Water (doh.wa.gov)"))

    header_cell = ws.cell(row=current_row, column=1, value="DISTRICT SUMMARY")
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = _header_fill("375623")
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row,   end_column=9
    )
    header_cell.alignment = Alignment(horizontal="left", vertical="center")
    current_row += 1

    for label, value in summary_items:
        bold = "TOTAL" in label
        lc   = ws.cell(row=current_row, column=1, value=label)
        vc   = ws.cell(row=current_row, column=2, value=value)
        lc.font = Font(bold=bold)
        vc.font = Font(bold=bold)
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(vertical="center")
        ws.merge_cells(
            start_row=current_row, start_column=2,
            end_row=current_row,   end_column=9
        )
        current_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    _set_col_widths(ws, {
        "A": 36, "B": 20, "C": 28, "D": 22,
        "E": 16, "F": 14, "G": 16, "H": 22, "I": 22,
    })


# ── Public API ────────────────────────────────────────────────────────────────

def _build_district_workbook(result: dict) -> tuple[Workbook, str]:
    if not result.get("found"):
        raise ValueError(
            f"District '{result.get('district_name')}' not found. Cannot generate report."
        )

    district_name = result["district_name"]
    summary = result["summary"]
    fixture_rows = result["fixture_rows"]
    detail_df = _build_detail_rows(fixture_rows)

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Grant Application")
    ws2 = wb.create_sheet("Full Detail")

    _write_grant_tab(ws1, detail_df, district_name, summary)
    _write_full_tab(ws2, detail_df, district_name, summary)

    filename = f"{_safe_filename(district_name)}_lead_report.xlsx"
    return wb, filename


def generate_district_xlsx_bytes(result: dict) -> tuple[bytes, str]:
    """
    Build a district report workbook in memory.

    Returns:
        (file_bytes, download_filename)
    """
    wb, filename = _build_district_workbook(result)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), filename


def generate_district_xlsx(result: dict, output_dir: str = "exports") -> str:
    """
    Generates a two-tab .xlsx report for a district.

    Args:
        result:     Dict returned by query_engine.search_district()
        output_dir: Directory to save the file (created if needed)

    Returns:
        Absolute path to the saved .xlsx file.
    """
    wb, filename = _build_district_workbook(result)

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    filepath = output_path / filename
    wb.save(filepath)

    return str(filepath.resolve())